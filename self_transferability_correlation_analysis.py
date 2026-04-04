import os
import torch
import torchvision.transforms.v2
from tqdm import tqdm, tqdm_notebook
import csv
from PIL import Image
import matplotlib.pyplot as plt
import json
import pandas as pd
from utils import *
import argparse
from transformations.edi import *
from transformations.transforms import *
from openpyxl import Workbook
import transformations.basic_transformations as bt

device = "cuda" if torch.cuda.is_available() else "cpu"

data_transform = transforms.Compose([
    transforms.Resize([224, 224]),
    transforms.ToTensor()])

mean, stddev = [0.485, 0.456, 0.406], [0.229, 0.224, 0.225]
COLUMNS = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V',
           'W', 'X', 'Y', 'Z']

model_name = 'resnet50'
surrogate = load_model(model_name)
surrogate = WrapperModel(surrogate, mean, stddev, False, surrogate).to(device).eval()

target_model_names = ['mobilenet_v2', 'efficientnet_b0', 'convnext', 'inception_v3', 'inception_v4_timm',
                      'inception_resnet_v2', 'xception', 'vit_base_patch16_224', 'swin', 'maxvit', 'twins_svt_base', 'pit', 'tnt', 'deit']

inceptions = ['xception', 'inception_v3', 'inception_resnet_v2', 'inception_v4_timm', 'tf_ens3_adv_inc_v3',
              'tf_ens4_adv_inc_v3', 'tf_ens_adv_inc_res_v2']

models = [WrapperModel(load_model(x), mean, stddev, True if x in inceptions else False, x).to(device).eval() for x in target_model_names]


def sample_mags(range):
    if len(range) == 0:
        return [0]
    elif len(range) == 2:
        return np.linspace(range[0], range[1], 100)
    elif len(range) == 4:
        x_samples = np.linspace(range[0], range[1], 10)
        y_samples = np.linspace(range[2], range[3], 10)
        xx, yy = np.meshgrid(x_samples, y_samples)
        points = np.c_[xx.ravel(), yy.ravel()]
        return points

trans_names = ['Rotation', 'Scaling', 'Shear', 'Perspective', 'Flip', 'Crop', 'Translate',
               'Solarize', 'Hue', 'Brightness', 'Contrast', 'Saturation']

trans_func = {'None': torchvision.transforms.v2.Identity(),
              'Rotation': lambda r_value: bt.rotation_trans(r=r_value),
              'Scaling': lambda r_value: bt.scaling_trans(r=r_value),
              'Crop': lambda r_value: bt.crop_trans(r=r_value),
              'Shear': lambda r_value: bt.shear_trans(r=r_value),
              'Perspective': lambda r_value: bt.perspective_trans(r=r_value),
              'Brightness': lambda r_value: bt.brightness_trans(r=r_value),
              'Saturation': lambda r_value: bt.saturation_trans(r=r_value),
              'Hue': lambda r_value: bt.hue_trans(r=r_value),
              'Solarize': lambda r_value: bt.solarize_trans(r=r_value),
              'Contrast': lambda r_value: bt.contrast_trans(r=r_value),
              'Flip': lambda r_value: bt.flip_trans(r=r_value),
              'Translate': lambda r_value: bt.translate_trans(r=r_value)
              }

trans_magitude = \
    {'Rotation': [0., 1.], 'Scaling': [0., 1.], 'Shear': [0., 1., 0., 1.], 'Perspective': [0., 1.],
     'Flip': [], 'Elastic': [0., 1.], 'Crop': [0., 1.], 'Translate': [0., 1., 0., 1.], 'Solarize': [0., 1.],
     'Hue': [0., 1.], 'Brightness': [0., 1.], 'Contrast': [0., 1.], 'Saturation': [0., 1.], 'Sharpeness': [0., 1.]}

wb = Workbook()
ws = wb.active
ws['B1'] = 'T_logit'

for i in range(len(trans_names)):
    column = COLUMNS[i + 5]
    ws[f'{column}1'] = trans_names[i]

for path_idx, adv_path in enumerate([
    'tab3_vanilla/resnet50_None_L-CE_Margin_Iter-900_Copies-1_Eps-16',
    'tab3_vanilla/resnet50_SI_L-CE_Margin_Iter-900_Copies-5_Eps-16',
    'tab3_vanilla/resnet50_Admix_L-CE_Margin_Iter-900_Copies-5_Eps-16',

    'extended_results_for_ST_analysis/resnet50_SSA_R-0.1',
    'extended_results_for_ST_analysis/resnet50_SSA_R-0.2',
    'extended_results_for_ST_analysis/resnet50_SSA_R-0.3',
    'extended_results_for_ST_analysis/resnet50_SSA_R-0.4',
    'tab3_vanilla/resnet50_SSA_L-CE_Margin_Iter-900_Copies-1_Eps-16',
    'extended_results_for_ST_analysis/resnet50_SSA_R-0.6',
    'extended_results_for_ST_analysis/resnet50_SSA_R-0.7',

    'tab3_vanilla/resnet50_RDI_L-CE_Margin_Iter-900_Copies-1_Eps-16',

    'extended_results_for_ST_analysis/resnet50_DI_R-1.02',
    'extended_results_for_ST_analysis/resnet50_DI_R-1.04',
    'extended_results_for_ST_analysis/resnet50_DI_R-1.06',
    'extended_results_for_ST_analysis/resnet50_DI_R-1.08',
    'tab3_vanilla/resnet50_DI_L-CE_Margin_Iter-900_Copies-1_Eps-16',
    'extended_results_for_ST_analysis/resnet50_DI_R-1.3',
    'extended_results_for_ST_analysis/resnet50_DI_R-1.5',
    'extended_results_for_ST_analysis/resnet50_DI_R-1.7',
    'extended_results_for_ST_analysis/resnet50_DI_R-1.9',
    'extended_results_for_ST_analysis/resnet50_DI_R-2.1',

    'tab3_vanilla/resnet50_BSR_L-CE_Margin_Iter-900_Copies-1_Eps-16',
    'tab3_vanilla/resnet50_ODI_L-CE_Margin_Iter-900_Copies-1_Eps-16',

    'extended_results_for_ST_analysis/resnet50_SIA_B-1',
    'extended_results_for_ST_analysis/resnet50_SIA_B-2',
    'tab3_vanilla/resnet50_SIA_L-CE_Margin_Iter-900_Copies-1_Eps-16',
    'extended_results_for_ST_analysis/resnet50_SIA_B-4',
    'extended_results_for_ST_analysis/resnet50_SIA_B-5',

    'tab3_vanilla/resnet50_TAug_L-CE_Margin_Iter-900_Copies-1_Eps-16',
    'tab3_vanilla/resnet50_HAug_L-CE_Margin_Iter-900_Copies-1_Eps-16',

    'extended_results_for_ST_analysis/resnet50_decowa_R-0.2',
    'extended_results_for_ST_analysis/resnet50_decowa_R-0.4',
    'extended_results_for_ST_analysis/resnet50_decowa_R-0.6',
    'extended_results_for_ST_analysis/resnet50_decowa_R-0.8',
    'extended_results_for_ST_analysis/resnet50_decowa_R-1.0',
    'extended_results_for_ST_analysis/resnet50_decowa_R-1.2',
    'extended_results_for_ST_analysis/resnet50_decowa_R-1.4',
    'extended_results_for_ST_analysis/resnet50_decowa_R-1.6',
    'extended_results_for_ST_analysis/resnet50_decowa_R-1.8',
    'tab3_vanilla/resnet50_decowa_L-CE_Margin_Iter-900_Copies-1_Eps-16',
    
]):
    print(adv_path)
    
    batch_size = 10
    image_id_list, label_ori_list, label_tar_list = load_ground_truth('./dataset/images.csv')
    input_path = './dataset/images/'
    adv_path = f'./results/{adv_path}/images/'
    num_batches = int(np.ceil(len(image_id_list) / batch_size))
    total_num = len(image_id_list)
    
    rank_per_trans = [0.] * len(trans_names)
    
    tar_logit_trans = [0.] * len(trans_names)
    tar_sigmoid_trans = [0.] * len(trans_names)
    
    victims_logit_cont = 0.  # [0.] * len(target_model_names)
    tar_prob = [0.] * len(target_model_names)
    
    for k in tqdm(range(0, num_batches)):
        batch_size_cur = min(batch_size, len(image_id_list) - k * batch_size)
        ori_X = torch.zeros(batch_size_cur, 3, 224, 224).cuda()
        adv_X = torch.zeros(batch_size_cur, 3, 224, 224).cuda()
        for i in range(batch_size_cur):
            ori_X[i] = data_transform(load_img(input_path + image_id_list[k * batch_size + i] + '.png'))
            adv_X[i] = data_transform(load_img(adv_path + image_id_list[k * batch_size + i] + '.png'))
        labels = torch.tensor(label_ori_list[k * batch_size:k * batch_size + batch_size_cur]).cuda()
        target_labels = torch.tensor(label_tar_list[k * batch_size:k * batch_size + batch_size_cur]).cuda()
        target_one_hot_labels = torch.zeros((target_labels.shape[0], 1000)).cuda()
        target_one_hot_labels.scatter_(1, target_labels.unsqueeze(1), 1)
        source_one_hot_labels = torch.zeros((labels.shape[0], 1000)).cuda()
        source_one_hot_labels.scatter_(1, labels.unsqueeze(1), 1)
        
        with torch.no_grad():
            for idx, name in enumerate(trans_names):
                magnitudes = sample_mags(trans_magitude[name])
                num_points = len(magnitudes)
                for mag in magnitudes:
                    trans_input = trans_func[name](mag)(torch.cat([ori_X, adv_X]))
                    num_trans = len(trans_input)
                    for nt in range(num_trans):
                        input = trans_input[nt].detach().clone()
                        clean_output = surrogate(input[0:batch_size_cur].clamp(0, 1)).softmax(1)
                        adv_output = surrogate(input[batch_size_cur:].clamp(0, 1)).softmax(1)
                        tar_logit_trans[idx] += ((adv_output - clean_output) * target_one_hot_labels).sum().item() / total_num / num_points / num_trans
            
            for i in range(len(target_model_names)):
                clean_output = models[i](ori_X).softmax(1)
                adv_output = models[i](adv_X).softmax(1)
                victims_logit_cont += ((adv_output - clean_output) * target_one_hot_labels).sum().item() / total_num / len(target_model_names)
    
    ws[f'A{path_idx + 2}'] = adv_path
    ws[f'B{path_idx + 2}'] = victims_logit_cont
    
    for i in range(len(trans_names)):
        column = COLUMNS[i + 5]
        ws[f'{column}{path_idx + 2}'] = tar_logit_trans[i]
    
    wb.save("./self_transferability_data.xlsx")


